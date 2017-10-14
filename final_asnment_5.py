import click
import MySQLdb
import openpyxl

db = MySQLdb.connect(host="localhost", user="root",
                     passwd="kandula123")
c = db.cursor()
#print("hello this is tarak")
global_db='eswar'
@click.group()
def cli():
    pass
@cli.command()
@click.argument('db_name',nargs=1)

def createdb(db_name):
    global global_db
    try:
        c.execute("CREATE DATABASE %s"%db_name)
    except:
        c.execute("DROP DATABASE %s"%db_name)
        c.execute("CREATE DATABASE %s"%db_name)

    c.execute("USE %s"%db_name)
    global_db=db_name
    sql = """CREATE TABLE students (
             NAME  CHAR(60) NOT NULL,
             COLLEGE  CHAR(60) NOT NULL,
             EMAIL_ID CHAR(80) NOT NULL,
             DBNAMES CHAR(20) NOT NULL,
             ID CHAR(40) NOT NULL,
             PRIMARY KEY (ID))"""


    c.execute(sql)
    sql = """CREATE TABLE marks (
             NO char(5),
             STUDENT_ID  CHAR(40) NOT NULL,
             PROBLEM1 CHAR(20),
             PROBLEM2 CHAR(20),
             PROBLEM3 CHAR(20),
             PROBLEM4 CHAR(20),
             TOTAL INT,
             ID CHAR(40) NOT NULL,
             FOREIGN KEY (ID) REFERENCES students(ID))"""


    c.execute(sql)



@cli.command()
@click.argument('dropdb_name',nargs=1)
def dropdb(dropdb_name):
    c.execute("DROP DATABASE %s" % dropdb_name)

@cli.command()
@click.argument('students_filename',nargs=1)
@click.argument('marks_filename',nargs=1)
def importdata(students_filename,marks_filename):
    c.execute('USE %s' %global_db)
    wb1 = openpyxl.load_workbook(students_filename)
    wk_sheet1= wb1.active
    wb2 = openpyxl.load_workbook(marks_filename)
    wk_sheet2 = wb2.active


    # openpyxl.worksheet.copier.WorksheetCopy('E:\MRND\students.xlsx', 'E:\MRND/teachers.xlsx')
    # target = wb.copy_worksheet(source)
    for i, row in enumerate(wk_sheet1.iter_rows()):
        if i==0:
            continue
        a=[]
        for j, col in enumerate(row):
            a.append(col.value)
            print(a[j])
        a.append(a[1].lower()+'_'+a[3].lower())
        sql = "INSERT INTO STUDENTS(NAME,COLLEGE,EMAIL_ID, DBNAMES, ID) VALUES ('%s', '%s', '%s', '%s', '%s' )" %\
              (a[0],a[1],a[2],a[3],a[4])
        c.execute(sql)
        print('tris')
    db.commit()
    l=list(wk_sheet2.iter_rows())
    print(len(l))
    for i, row in enumerate(l):
        if i==0:
            continue
        if row[0].value==None:
            break
        a = []
        for j, col in enumerate(row):
            a.append(col.value)
            print(a[j])
        a.append(a[1][7:-5])
        a[6]=int(a[6])
        print(a[1])
        print(a[7])

        sql = "INSERT INTO marks(NO,STUDENT_ID,PROBLEM1,PROBLEM2, PROBLEM3, PROBLEM4,TOTAL,ID) VALUES ('%s','%s', '%s', '%s','%s','%s', '%d', '%s' )" % (
            a[0], a[1], a[2], a[3], a[4],a[5],a[6],a[7])
        c.execute(sql)
    db.commit()


@cli.command()
def collegestats():
    c.execute("USE %s"%global_db)
    sql="SELECT students.COLLEGE,COUNT(marks.ID),MIN(marks.TOTAL),AVG(marks.TOTAL),MAX(marks.TOTAL)  FROM marks \
LEFT JOIN students ON marks.ID = students.ID GROUP BY COLLEGE"
    '''sql = "SELECT COLLEGE,COUNT(EMAIL_ID) FROM STUDENTS GROUP BY COLLEGE"'''
    # Fetch all the rows in a list of lists.
    c.execute(sql)
    results = c.fetchall()
    for row in results:
        college = row[0]
        count = row[1]
        min=row[2]
        avg=row[3]
        max=row[4]
        # Now print fetched result
        print('{} \n {} {} {} {}'.format(college,count,min,avg,max))

if __name__=='__main__':
    cli()




# Create table as per requirement


