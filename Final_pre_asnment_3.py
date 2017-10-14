import click
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter



@click.command()
@click.argument('input_html',nargs=1)
@click.argument('output_excel',nargs=1)
def cli(input_html,output_excel):
    wb1 = Workbook()
    wk_sheet_2 = wb1.active
    with open(input_html) as fp:
        soup = BeautifulSoup(fp, "html.parser")  # Parse the HTML as a string

    table = soup.find_all('table')[0]  # Grab the first table
    cell_dimension = 0
    for i, row in enumerate(table.find_all('tr')):
        if i == 0:
            for j, data in enumerate(row.find_all('th')):
                wk_sheet_2.column_dimensions[get_column_letter(j+1)].width=len(str(data.get_text()))
                wk_sheet_2.cell(row=i + 1, column=j + 1).font = Font(bold=True)
                wk_sheet_2.cell(row=i + 1, column=j + 1).value = data.get_text()
        else:
            for j, data in enumerate(row.find_all('td')):
                if j == 1:
                    if len(str(data.get_text())) > cell_dimension:
                        cell_dimension = len(str(data.get_text()))
                wk_sheet_2.cell(row=i + 1, column=j + 1).value = data.get_text()
    wk_sheet_2.column_dimensions['B'].width = cell_dimension
    wb1.save(output_excel)

if __name__ == '__main__':
    cli()

