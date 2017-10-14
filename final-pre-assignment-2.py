import click
import string

import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.styles.fonts import Font,Color
from copy import copy
from openpyxl.utils import get_column_letter


@click.command()
@click.option('--capitalize',is_flag=True,help='converts the string to capitalize')
@click.option('--preserve_styles',is_flag=True,help='copied along with along cell styles')
@click.argument('input',nargs=1)
@click.argument('output',nargs=1)
def cli(capitalize,preserve_styles,input,output):
    wb1 = openpyxl.load_workbook(input)
    print(input)
    wb2 = Workbook()
    l = wb1.get_sheet_names()
    print(output)
    dest_filename = output

    wk_sheet = wb1.get_sheet_by_name(l[0])
    wk_sheet_2 = wb2.active
    wk_sheet_2.title = l[0]
    print('=========',preserve_styles)
    if preserve_styles==True:
        for i in range(wk_sheet.max_column):
            wk_sheet_2.column_dimensions[get_column_letter(i + 1)].width = wk_sheet.column_dimensions[get_column_letter(i + 1)].width
    for i, row in enumerate(wk_sheet.iter_rows()):
        wk_sheet_2.row_dimensions[i + 1].height = wk_sheet.row_dimensions[i + 1].height
        for j, col in enumerate(row):
            if capitalize:
                wk_sheet_2.cell(row=i + 1, column=j + 1).value = string.capitalize(col.value)
            else:
                wk_sheet_2.cell(row=i + 1, column=j + 1).value = col.value

            if preserve_styles==True:
                wk_sheet_2.cell(row=i + 1, column=j + 1).font = copy(col.font)


    for k in range(1, len(l)):
        wk_sheet = wb1.get_sheet_by_name(l[k]);
        wk_sheet_2 = wb2.create_sheet(title=l[k])
        if preserve_styles==True:
            for i in range(wk_sheet.max_column):
                wk_sheet_2.column_dimensions[get_column_letter(i + 1)].width = wk_sheet.column_dimensions[get_column_letter(i + 1)].width
        for i, row in enumerate(wk_sheet.iter_rows()):
            wk_sheet_2.row_dimensions[i + 1].height = wk_sheet.row_dimensions[i + 1].height
            for j, col in enumerate(row):
                if capitalize:
                    wk_sheet_2.cell(row=i + 1, column=j + 1).value = string.capitalize(col.value)
                else:
                    wk_sheet_2.cell(row=i + 1, column=j + 1).value = col.value

                if preserve_styles==True:
                    wk_sheet_2.cell(row=i + 1, column=j + 1).font = copy(col.font)

    wb2.save(dest_filename)


if __name__ == '__main__':
    cli()






















