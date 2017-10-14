import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
from openpyxl.styles import Font
wb1 = openpyxl.load_workbook("input.xlsx")
wk_sheet_2 = wb1.active


#openpyxl.worksheet.copier.WorksheetCopy('E:\MRND\students.xlsx', 'E:\MRND/teachers.xlsx')
#target = wb.copy_worksheet(source)

html_string = '''
    <table>
          <tr>
              <td> Hello! </td>
              <td> Table </td>
          </tr>
          <tr>
              <td> Tarak</td>
              <td> ram </td>
          </tr>
          <tr>
              <td> eswar </td>
              <td> kandula </td>
          </tr>
      </table>
  '''
with open("html_input.html") as fp:

    soup = BeautifulSoup(fp,"html.parser") #Parse the HTML as a string

table = soup.find_all('table')[0]  # Grab the first table

#new_table = pd.DataFrame(columns=range(0, 2), index=[0])  # I know the size

'''for i,row in enumerate(wk_sheet_2.iter_rows()):
  for j,col in enumerate(row):



'''
#wk_sheet_2['C1'].style.font.bold=True
cell_dimension=0
for i,row in enumerate(table.find_all('tr')):
    if i==0:
        for j, data in enumerate(row.find_all('th')):
            wk_sheet_2.cell(row=i + 1, column=j + 1).font = Font(bold=True)
            wk_sheet_2.cell(row=i + 1, column=j + 1).value = data.get_text()
            #wk_sheet_2.cell(row=i + 1, column=j + 1).font.bold=True
            print(i, j)
            print(data.get_text())
    else:
        for j, data in enumerate(row.find_all('td')):
            if j==1:
                if len(str(data.get_text()))> cell_dimension:
                    cell_dimension=len(str(data.get_text()))
            #print(len(str(data.get_text())))
            wk_sheet_2.cell(row=i + 1, column=j + 1).value = data.get_text()
            print(i, j)
            print(data.get_text())
wk_sheet_2.column_dimensions['B'].width=cell_dimension
print(wk_sheet_2[1])
#wk_sheet_2[1].font.bold=True
wb1.save('input.xlsx')