import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl import Workbook
from openpyxl.styles.fonts import Font,Color
from copy import copy
from openpyxl.utils import get_column_letter
def copy_style(src_cell, dest_cell):
    dest_cell.font = src_cell.font
    dest_cell.fill = src_cell.fill
    dest_cell.border = src_cell.border
    dest_cell.alignment = src_cell.alignment
    dest_cell.number_format = src_cell.number_format

wb1 = openpyxl.load_workbook('students.xlsx')
wb2 = Workbook()
#wb2 = openpyxl.load_workbook('students.xlsx')
#source = wb.active

l=wb1.get_sheet_names()
print(l)
dest_filename = 'output.xlsx'
wk_sheet = wb1.get_sheet_by_name(l[0])
type(wk_sheet)
wk_sheet_2 = wb2.active
wk_sheet_2.title = l[0]

# wk_sheet_2 = wb2.get_sheet_by_name(sheet);

for i in range(wk_sheet.max_column):
    wk_sheet_2.column_dimensions[get_column_letter(i + 1)].width = wk_sheet.column_dimensions[get_column_letter(i + 1)].width
for i, row in enumerate(wk_sheet.iter_rows()):
    print(i,row)
    wk_sheet_2.row_dimensions[i + 1].height = wk_sheet.row_dimensions[i + 1].height
    #wk_sheet_2.column_dimensions['A'].width=wk_sheet.column_dimensions['A'].width

    for j, col in enumerate(row):
        print(j,col)

        wk_sheet_2.cell(row=i + 1, column=j + 1).value = col.value
        wk_sheet_2.cell(row=i + 1, column=j + 1).font = copy(col.font)

for k in range(1,len(l)):
    wk_sheet=wb1.get_sheet_by_name(l[k]);
    type(wk_sheet)
    wk_sheet_2 = wb2.create_sheet(title=l[k])
    #wk_sheet_2 = wb2.get_sheet_by_name(sheet);
    for i in range(wk_sheet.max_column):
        wk_sheet_2.column_dimensions[get_column_letter(i+1)].width=wk_sheet.column_dimensions[get_column_letter(i+1)].width
    for i, row in enumerate(wk_sheet.iter_rows()):
        wk_sheet_2.row_dimensions[i+1].height=wk_sheet.row_dimensions[i + 1].height
        for j, col in enumerate(row):
            wk_sheet_2.cell(row=i + 1, column=j + 1).value = col.value
            wk_sheet_2.cell(row=i + 1, column=j + 1).font =copy(col.font)

            '''
            for row in range(1, 40):
                ws1.append(range(600))

            ws2 = wb.create_sheet(title="Pi")
            ws2['F5'] = 3.14

for i in l:
    print(i)
print(l)

sheet1 = wb.get_sheet_by_name('Current')
type(sheet1)
anotherSheet = wb.active
sheet2 = wb.get_sheet_by_name('Deletions')
type(sheet2)

#openpyxl.worksheet.copier.WorksheetCopy('E:\MRND\students.xlsx', 'E:\MRND/teachers.xlsx')
#target = wb.copy_worksheet(source)
for i,row in enumerate(sheet1.iter_rows()):
  for j,col in enumerate(row):
    sheet2.cell(row=i+1,column=j+1).value = col.value

wb.save('E:\MRND/students.xlsx')


====================================

from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'teachers.xlsx'
ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
     ws1.append(range(600))

ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14

ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
'''



wb2.save(dest_filename)